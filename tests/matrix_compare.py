"""Compare two module 2 matrix CSV/XLSX files and render a visual diff report."""

from __future__ import annotations

import argparse
import csv
import html
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence


@dataclass(frozen=True)
class MatrixData:
    row_label: str
    columns: List[str]
    rows: List[str]
    values: Dict[str, Dict[str, float]]


@dataclass(frozen=True)
class MatrixComparison:
    matches: bool
    differing_cells: int
    compared_cells: int
    max_absolute_difference: float
    missing_rows: List[str]
    unexpected_rows: List[str]
    missing_columns: List[str]
    unexpected_columns: List[str]
    diff_csv: Path
    html_report: Path

    def summary(self) -> str:
        status = "MATCH" if self.matches else "DIFFERENT"
        return (
            f"{status}: {self.differing_cells}/{self.compared_cells} shared cells differ; "
            f"max absolute difference={self.max_absolute_difference:.6g}; "
            f"rows missing/unexpected={len(self.missing_rows)}/{len(self.unexpected_rows)}; "
            f"columns missing/unexpected={len(self.missing_columns)}/{len(self.unexpected_columns)}"
        )


def _parse_matrix_rows(path: Path, source_rows: Iterable[Sequence[object]]) -> MatrixData:
    iterator = iter(source_rows)
    try:
        raw_header = list(next(iterator))
    except StopIteration as exc:
        raise ValueError(f"{path} is empty") from exc

    while raw_header and raw_header[-1] in (None, ""):
        raw_header.pop()
    if len(raw_header) < 2:
        raise ValueError(f"{path} must contain a row-label column and at least one sample column")

    row_label = str(raw_header[0]) if raw_header[0] not in (None, "") else "kmer"
    columns = [str(value) for value in raw_header[1:]]
    if any(not column for column in columns):
        raise ValueError(f"{path} contains an empty sample column name")
    if len(columns) != len(set(columns)):
        raise ValueError(f"{path} contains duplicate sample columns")

    rows: List[str] = []
    values: Dict[str, Dict[str, float]] = {}
    for line_number, source_row in enumerate(iterator, start=2):
        row = list(source_row)
        if not row or not any(value not in (None, "") for value in row):
            continue
        extra_values = row[len(raw_header) :]
        if any(value not in (None, "") for value in extra_values):
            raise ValueError(
                f"{path}:{line_number} has values beyond the {len(raw_header)} matrix columns"
            )
        row = row[: len(raw_header)]
        if len(row) != len(raw_header):
            raise ValueError(
                f"{path}:{line_number} has {len(row)} fields; expected {len(raw_header)}"
            )
        if row[0] in (None, ""):
            raise ValueError(f"{path}:{line_number} contains an empty row label")
        row_name = str(row[0])
        if row_name in values:
            raise ValueError(f"{path}:{line_number} contains duplicate row {row_name!r}")
        try:
            numeric_values = [float(value) for value in row[1:]]
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{path}:{line_number} contains a non-numeric matrix value") from exc
        if not all(math.isfinite(value) for value in numeric_values):
            raise ValueError(f"{path}:{line_number} contains NaN or an infinite value")
        rows.append(row_name)
        values[row_name] = dict(zip(columns, numeric_values))

    return MatrixData(row_label=row_label, columns=columns, rows=rows, values=values)


def read_matrix(path: Path) -> MatrixData:
    """Read a matrix from CSV or from the active worksheet of an XLSX workbook."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return _parse_matrix_rows(path, csv.reader(handle))
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "XLSX comparison requires openpyxl; install backend/requirements.txt"
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            return _parse_matrix_rows(path, worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()
    raise ValueError(f"{path} has unsupported format {suffix!r}; use .csv or .xlsx")


def _ordered_union(first: Sequence[str], second: Sequence[str]) -> List[str]:
    first_values = set(first)
    return list(first) + [value for value in second if value not in first_values]


def _cell_color(difference: float, scale: float) -> str:
    if difference == 0 or scale == 0:
        return "#ffffff"
    intensity = min(abs(difference) / scale, 1.0)
    channel = round(255 - 120 * intensity)
    if difference > 0:
        return f"#ff{channel:02x}{channel:02x}"
    return f"#{channel:02x}{channel:02x}ff"


def compare_matrix_files(
    expected_path: Path,
    actual_path: Path,
    output_dir: Path,
    *,
    absolute_tolerance: float = 1e-12,
    relative_tolerance: float = 1e-9,
    max_visual_rows: int = 100,
) -> MatrixComparison:
    """Compare matrices by labels, write a complete CSV diff, and render an HTML heatmap."""
    expected = read_matrix(Path(expected_path))
    actual = read_matrix(Path(actual_path))
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    missing_rows = [row for row in expected.rows if row not in actual.values]
    unexpected_rows = [row for row in actual.rows if row not in expected.values]
    missing_columns = [column for column in expected.columns if column not in actual.columns]
    unexpected_columns = [column for column in actual.columns if column not in expected.columns]
    all_rows = _ordered_union(expected.rows, actual.rows)
    all_columns = _ordered_union(expected.columns, actual.columns)
    shared_rows = [row for row in expected.rows if row in actual.values]
    shared_columns = [column for column in expected.columns if column in actual.columns]

    differences: Dict[str, Dict[str, float]] = {}
    differing_cells = 0
    max_absolute_difference = 0.0
    for row in shared_rows:
        differences[row] = {}
        for column in shared_columns:
            expected_value = expected.values[row][column]
            actual_value = actual.values[row][column]
            difference = actual_value - expected_value
            differences[row][column] = difference
            max_absolute_difference = max(max_absolute_difference, abs(difference))
            if not math.isclose(
                actual_value,
                expected_value,
                abs_tol=absolute_tolerance,
                rel_tol=relative_tolerance,
            ):
                differing_cells += 1

    diff_csv = output_dir / "matrix_diff.csv"
    with diff_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow([expected.row_label, *all_columns])
        for row in all_rows:
            writer.writerow(
                [
                    row,
                    *[
                        differences.get(row, {}).get(column, "")
                        for column in all_columns
                    ],
                ]
            )

    row_scores = {
        row: max((abs(value) for value in differences.get(row, {}).values()), default=0.0)
        for row in all_rows
    }
    visual_rows = sorted(
        all_rows,
        key=lambda row: (
            row not in expected.values or row not in actual.values,
            row_scores[row],
        ),
        reverse=True,
    )[:max_visual_rows]
    scale = max_absolute_difference

    heatmap_rows = []
    for row in visual_rows:
        cells = [f"<th>{html.escape(row)}</th>"]
        for column in all_columns:
            if row not in expected.values or row not in actual.values:
                cells.append('<td class="missing" title="Row is absent from one matrix">missing</td>')
            elif column not in expected.columns or column not in actual.columns:
                cells.append('<td class="missing" title="Column is absent from one matrix">missing</td>')
            else:
                expected_value = expected.values[row][column]
                actual_value = actual.values[row][column]
                difference = actual_value - expected_value
                color = _cell_color(difference, scale)
                tooltip = (
                    f"expected={expected_value:.8g}, actual={actual_value:.8g}, "
                    f"actual-expected={difference:.8g}"
                )
                cells.append(
                    f'<td style="background:{color}" title="{html.escape(tooltip)}">'
                    f"{difference:.3g}</td>"
                )
        heatmap_rows.append(f"<tr>{''.join(cells)}</tr>")

    matches = not (
        differing_cells
        or missing_rows
        or unexpected_rows
        or missing_columns
        or unexpected_columns
    )
    comparison = MatrixComparison(
        matches=matches,
        differing_cells=differing_cells,
        compared_cells=len(shared_rows) * len(shared_columns),
        max_absolute_difference=max_absolute_difference,
        missing_rows=missing_rows,
        unexpected_rows=unexpected_rows,
        missing_columns=missing_columns,
        unexpected_columns=unexpected_columns,
        diff_csv=diff_csv,
        html_report=output_dir / "matrix_diff.html",
    )

    def joined(values: Sequence[str]) -> str:
        if not values:
            return "none"
        displayed = ", ".join(values[:50])
        if len(values) > 50:
            displayed += f" … (+{len(values) - 50} more; see matrix_diff.csv)"
        return html.escape(displayed)

    report = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Module 2 matrix comparison</title>
<style>
body {{ font: 14px system-ui, sans-serif; margin: 2rem; color: #172033; }}
.status {{ padding: .8rem 1rem; border-radius: 6px; background: {"#e5f7eb" if matches else "#fff0f0"}; }}
table {{ border-collapse: collapse; margin-top: 1rem; }}
th, td {{ border: 1px solid #ccd2dc; padding: .35rem .5rem; text-align: right; white-space: nowrap; }}
th {{ background: #f3f5f8; }}
.missing {{ background: #d8dbe2; color: #555; }}
.legend span {{ display: inline-block; padding: .25rem .6rem; margin-right: .3rem; }}
</style>
</head>
<body>
<h1>Module 2 matrix comparison</h1>
<p class="status"><strong>{comparison.summary()}</strong></p>
<p>Expected: {html.escape(str(Path(expected_path)))}</p>
<p>Actual: {html.escape(str(Path(actual_path)))}</p>
<ul>
<li>Missing rows: {joined(missing_rows)}</li>
<li>Unexpected rows: {joined(unexpected_rows)}</li>
<li>Missing columns: {joined(missing_columns)}</li>
<li>Unexpected columns: {joined(unexpected_columns)}</li>
</ul>
<p class="legend"><span style="background:#8787ff">actual is lower</span>
<span style="background:#ff8787">actual is higher</span>
<span style="background:#d8dbe2">row/column missing</span></p>
<p>Cells show <code>actual - expected</code>. Hover for both values.
Showing {len(visual_rows)} of {len(all_rows)} rows, ordered by largest difference.</p>
<table>
<thead><tr><th>{html.escape(expected.row_label)}</th>{''.join(f"<th>{html.escape(column)}</th>" for column in all_columns)}</tr></thead>
<tbody>{''.join(heatmap_rows)}</tbody>
</table>
</body>
</html>
"""
    comparison.html_report.write_text(report, encoding="utf-8")
    return comparison


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("expected", type=Path, help="Approved/baseline matrix (.csv or .xlsx)")
    parser.add_argument("actual", type=Path, help="Newly generated matrix (.csv or .xlsx)")
    parser.add_argument("-o", "--output-dir", type=Path, default=Path("test-artifacts/module2-matrix"))
    parser.add_argument("--atol", type=float, default=1e-12, help="Absolute numeric tolerance")
    parser.add_argument("--rtol", type=float, default=1e-9, help="Relative numeric tolerance")
    args = parser.parse_args()

    comparison = compare_matrix_files(
        args.expected,
        args.actual,
        args.output_dir,
        absolute_tolerance=args.atol,
        relative_tolerance=args.rtol,
    )
    print(comparison.summary())
    print(f"CSV diff: {comparison.diff_csv}")
    print(f"Visual report: {comparison.html_report}")
    return 0 if comparison.matches else 1


if __name__ == "__main__":
    raise SystemExit(main())
